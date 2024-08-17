from datetime import datetime
import sys
import shutil
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QTabWidget, QWidget, QVBoxLayout, QFormLayout,
    QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem, QMessageBox,
    QComboBox, QHBoxLayout, QDateEdit, QFileDialog
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPalette, QColor
from openpyxl import Workbook, load_workbook

class LibraryManagementSystem(QMainWindow):
    def __init__(self):
        super().__init__()
        self.excel_file = 'library_data.xlsx'
        self.backup_file = 'library_data_backup.xlsx'
        self.logged_in = False
        self.current_student_id = None
        self.wb = None
        self.initUI()

    def initUI(self):
        self.setStyleSheet("background-color: #f5f5f5;")
        self.create_or_load_library_data_excel()
        self.show_login_window()

    def create_or_load_library_data_excel(self):
        try:
            self.wb = load_workbook(self.excel_file)
            print(f"Loaded existing Library Data Excel file: {self.excel_file}")
        except FileNotFoundError:
            print(f"Library Data Excel file not found at: {self.excel_file}")
            self.create_library_data_excel()

    def create_library_data_excel(self):
        self.wb = Workbook()
        ws_admin = self.wb.create_sheet(title='Admin')
        ws_students = self.wb.create_sheet(title='students')
        ws_books = self.wb.create_sheet(title='Books')
        ws_borrowing = self.wb.create_sheet(title='Borrowing')
        ws_returned = self.wb.create_sheet(title='Returned')

        ws_admin.append(['Username', 'Password', 'Role'])
        ws_admin.append(['admin', '123', 'admin'])
        ws_students.append(['ID', 'student ID', 'Position', 'student Name', 'Borrowed Count'])
        ws_books.append(['ID', 'Book ID', 'Book Title', 'Author', 'Copies Available', 'Borrowed Count'])
        ws_borrowing.append(['ID', 'Book ID', 'Book Title', 'student ID', 'student Name', 'Borrowed Date'])
        ws_returned.append(['ID', 'Book ID', 'Book Title', 'student ID', 'student Name', 'Borrowed Date', 'Returned Date'])

        self.wb.save(self.excel_file)
        print(f"Created new Library Data Excel file with schema: {self.excel_file}")

    def show_login_window(self):
        self.login_window = QWidget()
        self.login_window.setWindowTitle("تسجيل الدخول القائد")
        self.login_window.setGeometry(100, 100, 300, 250)

        layout = QFormLayout()

        self.username_input = QLineEdit()
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)

        login_button = QPushButton("تسجيل الدخول")
        login_button.clicked.connect(self.login)

        layout.addRow(QLabel("اسم المستخدم:"), self.username_input)
        layout.addRow(QLabel("كلمة السر:"), self.password_input)
        layout.addWidget(login_button)

        self.login_window.setLayout(layout)
        self.login_window.show()

    def login(self):
        username = self.username_input.text()
        password = self.password_input.text()

        admin_sheet = self.wb['Admin']
        for row in admin_sheet.iter_rows(min_row=2, values_only=True):
            if username == row[0] and password == row[1]:
                self.logged_in = True
                self.current_student_id = username
                QMessageBox.information(self, "تم تسجيل الدخول", f"مرحبا, {username}!")
                self.create_main_window()
                self.login_window.close()
                return

        QMessageBox.critical(self, "فشل تسجيل الدخول", "خطاء فالاسم او كلمة السر.")

    def create_main_window(self):
        self.main_window = QMainWindow()
        self.main_window.setWindowTitle("نظام المكتبه الذكى")
        self.main_window.setGeometry(100, 100, 1200, 800)

        self.tabs = QTabWidget()
        self.tabs.addTab(self.create_students_tab(), "الطلاب")
        self.tabs.addTab(self.create_books_tab(), "الكتب")
        self.tabs.addTab(self.create_borrowing_tab(), "عمليات الاستعاره")
        self.tabs.addTab(self.create_returned_tab(), "عمليات الاعاده")
        self.tabs.addTab(self.create_reports_tab(), "التقارير")

        backup_button = QPushButton("نسخة احتياطيه من البيانات")
        backup_button.clicked.connect(self.backup_data)

        restore_button = QPushButton("استعادة اخر نسخه احتياطيه")
        restore_button.clicked.connect(self.restore_data)

        layout = QVBoxLayout()
        layout.addWidget(self.tabs)
        layout.addWidget(backup_button)
        layout.addWidget(restore_button)

        central_widget = QWidget()
        central_widget.setLayout(layout)
        self.main_window.setCentralWidget(central_widget)
        self.main_window.show()

    def create_students_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        self.students_table = QTableWidget()
        self.students_table.setColumnCount(5)
        self.students_table.setHorizontalHeaderLabels(["ID", "الرقم العسكرى", "الرتبه", "الاسم", "عدد مرات الاستعاره"])
        layout.addWidget(self.students_table)

        self.load_students_from_excel()

        add_button = QPushButton("اضافة طالب جديد")
        add_button.clicked.connect(self.add_student)
        layout.addWidget(add_button)

        search_layout = QHBoxLayout()
        self.search_student_input = QLineEdit()
        self.search_student_input.setPlaceholderText("بحث طلاب...")
        self.search_student_input.textChanged.connect(self.search_students)
        search_layout.addWidget(self.search_student_input)
        layout.addLayout(search_layout)

        tab.setLayout(layout)
        return tab

    def create_books_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        self.books_table = QTableWidget()
        self.books_table.setColumnCount(6)
        self.books_table.setHorizontalHeaderLabels(["ID", "رقم الكتاب", "اسم الكتاب", "الكاتب", "عدد النسخ المتاحه", "عدد مرات الاستعاره"])
        layout.addWidget(self.books_table)

        self.load_books_from_excel()

        add_button = QPushButton("اضافة كتاب")
        add_button.clicked.connect(self.add_book)
        layout.addWidget(add_button)

        search_layout = QHBoxLayout()
        self.search_book_input = QLineEdit()
        self.search_book_input.setPlaceholderText("بحث كتب...")
        self.search_book_input.textChanged.connect(self.search_books)
        search_layout.addWidget(self.search_book_input)
        layout.addLayout(search_layout)

        tab.setLayout(layout)
        return tab

    def create_borrowing_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        self.borrowing_table = QTableWidget()
        self.borrowing_table.setColumnCount(6)
        self.borrowing_table.setHorizontalHeaderLabels(["ID", "رقم الكتاب", "اسم الكتاب", "الرقم العسكرى", "اسم الطالب", "تاريخ الاستعاره"])
        layout.addWidget(self.borrowing_table)

        self.load_borrowing_from_excel()

        borrow_button = QPushButton("استعاره جديده")
        borrow_button.clicked.connect(self.borrow_book)
        layout.addWidget(borrow_button)

        tab.setLayout(layout)
        return tab

    def create_returned_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        self.returned_table = QTableWidget()
        self.returned_table.setColumnCount(7)
        self.returned_table.setHorizontalHeaderLabels(["ID", "رقم الكتاب", "اسم الكتاب", "الرقم العسكرى", "اسم الطالب", "تاريخ الاستعاره", "تاريخ الاستعاده"])
        layout.addWidget(self.returned_table)

        self.load_returned_from_excel()

        return_button = QPushButton("استعادة كتاب")
        return_button.clicked.connect(self.return_book)
        layout.addWidget(return_button)

        tab.setLayout(layout)
        return tab

    def create_reports_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        self.reports_table = QTableWidget()
        self.reports_table.setColumnCount(7)
        self.reports_table.setHorizontalHeaderLabels(["ID", "رقم الكتاب", "اسم الكتاب", "الرقم العسكرى", "اسم الطالب", "تاريخ الاستعاره", "تاريخ الاستعاده"])
        layout.addWidget(self.reports_table)

        generate_report_button = QPushButton("انشاء تقرير")
        generate_report_button.clicked.connect(self.generate_report)
        layout.addWidget(generate_report_button)

        tab.setLayout(layout)
        return tab

    def load_students_from_excel(self):
        ws_students = self.wb['students']
        self.students_table.setRowCount(ws_students.max_row - 1)
        for i, row in enumerate(ws_students.iter_rows(min_row=2, values_only=True)):
            for j, value in enumerate(row):
                self.students_table.setItem(i, j, QTableWidgetItem(str(value)))

    def load_books_from_excel(self):
        ws_books = self.wb['Books']
        self.books_table.setRowCount(ws_books.max_row - 1)
        for i, row in enumerate(ws_books.iter_rows(min_row=2, values_only=True)):
            for j, value in enumerate(row):
                self.books_table.setItem(i, j, QTableWidgetItem(str(value)))

    def load_borrowing_from_excel(self):
        ws_borrowing = self.wb['Borrowing']
        self.borrowing_table.setRowCount(ws_borrowing.max_row - 1)
        for i, row in enumerate(ws_borrowing.iter_rows(min_row=2, values_only=True)):
            for j, value in enumerate(row):
                self.borrowing_table.setItem(i, j, QTableWidgetItem(str(value)))

    def load_returned_from_excel(self):
        ws_returned = self.wb['Returned']
        self.returned_table.setRowCount(ws_returned.max_row - 1)
        for i, row in enumerate(ws_returned.iter_rows(min_row=2, values_only=True)):
            for j, value in enumerate(row):
                self.returned_table.setItem(i, j, QTableWidgetItem(str(value)))

    def add_student(self):
        self.add_student_window = QWidget()
        self.add_student_window.setWindowTitle("اضافة طالب جديد")
        self.add_student_window.setGeometry(100, 100, 400, 300)

        layout = QFormLayout()

        self.add_student_id = QLineEdit()
        self.add_student_name = QLineEdit()
        self.add_student_position = QLineEdit()

        add_button = QPushButton("اضافة طالب جديد")
        add_button.clicked.connect(self.save_student)

        layout.addRow(QLabel("الرقم العسكرى:"), self.add_student_id)
        layout.addRow(QLabel("الاسم:"), self.add_student_name)
        layout.addRow(QLabel("الرتبه:"), self.add_student_position)
        layout.addWidget(add_button)

        self.add_student_window.setLayout(layout)
        self.add_student_window.show()

    def save_student(self):
        student_id = self.add_student_id.text()
        student_name = self.add_student_name.text()
        student_position = self.add_student_position.text()

        if not student_id or not student_name or not student_position:
            QMessageBox.warning(self, "خطأ", "برجاء ملئ جميع المدخلات.")
            return

        ws_students = self.wb['students']
        new_id = ws_students.max_row
        ws_students.append([new_id, student_id, student_position, student_name, 0])
        self.wb.save(self.excel_file)
        self.load_students_from_excel()
        self.add_student_window.close()

    def add_book(self):
        self.add_book_window = QWidget()
        self.add_book_window.setWindowTitle("اضافة كتاب جديد")
        self.add_book_window.setGeometry(100, 100, 400, 300)

        layout = QFormLayout()

        self.add_book_id = QLineEdit()
        self.add_book_title = QLineEdit()
        self.add_book_author = QLineEdit()
        self.add_book_copies = QLineEdit()

        add_button = QPushButton("اضافة كتاب جديد")
        add_button.clicked.connect(self.save_book)

        layout.addRow(QLabel("رقم الكتاب:"), self.add_book_id)
        layout.addRow(QLabel("اسم الكتاب:"), self.add_book_title)
        layout.addRow(QLabel("الكاتب:"), self.add_book_author)
        layout.addRow(QLabel("عدد النسخ:"), self.add_book_copies)
        layout.addWidget(add_button)

        self.add_book_window.setLayout(layout)
        self.add_book_window.show()

    def save_book(self):
        book_id = self.add_book_id.text()
        book_title = self.add_book_title.text()
        book_author = self.add_book_author.text()
        book_copies = self.add_book_copies.text()

        if not book_id or not book_title or not book_author or not book_copies:
            QMessageBox.warning(self,"خطأ", "برجاء ملئ جميع المدخلات.")
            return

        ws_books = self.wb['Books']
        new_id = ws_books.max_row
        ws_books.append([new_id, book_id, book_title, book_author, int(book_copies), 0])
        self.wb.save(self.excel_file)
        self.load_books_from_excel()
        self.add_book_window.close()

    def borrow_book(self):
        self.borrow_book_window = QWidget()
        self.borrow_book_window.setWindowTitle("استعارة كتاب")
        self.borrow_book_window.setGeometry(100, 100, 400, 300)

        layout = QFormLayout()

        self.borrow_book_dropdown = QComboBox()
        self.borrow_student_dropdown = QComboBox()
        self.borrow_date = QDateEdit()
        self.borrow_date.setDate(datetime.now())

        # Load books into the dropdown
        ws_books = self.wb['Books']
        for row in ws_books.iter_rows(min_row=2, values_only=True):
            self.borrow_book_dropdown.addItem(row[2], userData=row[1])

        # Load students into the dropdown
        ws_students = self.wb['students']
        for row in ws_students.iter_rows(min_row=2, values_only=True):
            self.borrow_student_dropdown.addItem(row[3], userData=row[1])

        borrow_button = QPushButton("استعارة كتاب")
        borrow_button.clicked.connect(self.save_borrowing)

        layout.addRow(QLabel("الكتاب:"), self.borrow_book_dropdown)
        layout.addRow(QLabel("الطالب:"), self.borrow_student_dropdown)
        layout.addRow(QLabel("التاريخ:"), self.borrow_date)
        layout.addWidget(borrow_button)

        self.borrow_book_window.setLayout(layout)
        self.borrow_book_window.show()

    def save_borrowing(self):
        book_id = self.borrow_book_dropdown.currentData()
        student_id = self.borrow_student_dropdown.currentData()
        borrow_date = self.borrow_date.date().toString(Qt.ISODate)

        if not book_id or not student_id:
            QMessageBox.warning(self, "خطأ", "برجاء اختيار كلا من الطالب والكتاب.")
            return

        ws_books = self.wb['Books']
        for row in ws_books.iter_rows(min_row=2, values_only=False):
            if row[1].value == book_id:
                if row[4].value <= 0:
                    QMessageBox.warning(self, "الكتاب غير متاح", "لم يتبقى نسخ من الكتاب.")
                    return
                row[4].value -= 1
                row[5].value += 1

        ws_borrowing = self.wb['Borrowing']
        new_id = ws_borrowing.max_row
        book_title = self.borrow_book_dropdown.currentText()
        student_name = self.borrow_student_dropdown.currentText()
        ws_borrowing.append([new_id, book_id, book_title, student_id, student_name, borrow_date])
        self.wb.save(self.excel_file)
        self.load_borrowing_from_excel()
        self.borrow_book_window.close()

    def return_book(self):
        self.return_book_window = QWidget()
        self.return_book_window.setWindowTitle("استعادة الكتاب")
        self.return_book_window.setGeometry(100, 100, 400, 300)

        layout = QFormLayout()

        self.return_book_dropdown = QComboBox()
        self.return_student_dropdown = QComboBox()
        self.return_date = QDateEdit()
        self.return_date.setDate(datetime.now())

        # Load borrowed books into the dropdown
        ws_borrowing = self.wb['Borrowing']
        borrowed_books = set()
        borrowed_students = set()
        
        for row in ws_borrowing.iter_rows(min_row=2, values_only=True):
            borrowed_books.add((row[1], row[2]))  # (Book ID, Book Title)
            borrowed_students.add((row[3], row[4]))  # (student ID, student Name)
        
        # Populate the book dropdown with borrowed books
        for book_id, book_title in borrowed_books:
            self.return_book_dropdown.addItem(book_title, userData=book_id)

        # Populate the student dropdown with borrowed students
        for student_id, student_name in borrowed_students:
            self.return_student_dropdown.addItem(student_name, userData=student_id)

        return_button = QPushButton("استعادة كتاب")
        return_button.clicked.connect(self.save_returning)

        layout.addRow(QLabel("الكتاب:"), self.return_book_dropdown)
        layout.addRow(QLabel("الطالب:"), self.return_student_dropdown)
        layout.addRow(QLabel("تاريخ الاستعاده:"), self.return_date)
        layout.addWidget(return_button)

        self.return_book_window.setLayout(layout)
        self.return_book_window.show()

    def save_returning(self):
        book_id = self.return_book_dropdown.currentData()
        student_id = self.return_student_dropdown.currentData()
        return_date = self.return_date.date().toString(Qt.ISODate)

        if not book_id or not student_id:
            QMessageBox.warning(self, "خطأ", "برجاء اختيار كلا من الطالب والكتاب.")
            return

        ws_borrowing = self.wb['Borrowing']
        ws_returned = self.wb['Returned']
        borrowed_date = None

        # Find and remove the borrowing record
        for row in ws_borrowing.iter_rows(min_row=2, values_only=False):
            if row[1].value == book_id and row[3].value == student_id:
                borrowed_date = row[5].value
                ws_borrowing.delete_rows(row[0].row)
                break

        if borrowed_date is None:
            QMessageBox.warning(self, "خطأ فالاستعاده", "لا يوجد استعاره لهذا الكتاب او الطالب")
            return

        # Update the "Books" sheet to increase the available copies
        ws_books = self.wb['Books']
        for row in ws_books.iter_rows(min_row=2, values_only=False):
            if row[1].value == book_id:
                row[4].value += 1
                break

        # Add a record to the "Returned" sheet
        new_id = ws_returned.max_row
        book_title = self.return_book_dropdown.currentText()
        student_name = self.return_student_dropdown.currentText()
        ws_returned.append([new_id, book_id, book_title, student_id, student_name, borrowed_date, return_date])
        
        # Save the changes
        self.wb.save(self.excel_file)
        self.load_returned_from_excel()
        self.return_book_window.close()

    def search_books(self):
        search_text = self.search_book_input.text().lower()
        ws_books = self.wb['Books']
        self.books_table.setRowCount(0)
        for row in ws_books.iter_rows(min_row=2, values_only=True):
            if search_text in row[2].lower():
                self.books_table.insertRow(self.books_table.rowCount())
                for i, value in enumerate(row):
                    self.books_table.setItem(self.books_table.rowCount() - 1, i, QTableWidgetItem(str(value)))

    def search_students(self):
        search_text = self.search_student_input.text().lower()
        ws_students = self.wb['students']
        self.students_table.setRowCount(0)
        for row in ws_students.iter_rows(min_row=2, values_only=True):
            if search_text in row[3].lower():
                self.students_table.insertRow(self.students_table.rowCount())
                for i, value in enumerate(row):
                    self.students_table.setItem(self.students_table.rowCount() - 1, i, QTableWidgetItem(str(value)))

    def backup_data(self):
        shutil.copy(self.excel_file, self.backup_file)
        QMessageBox.information(self, "النسخة الاحتياطيه", "تم اخذ نسخه احتياطيه بنجاح.")

    def restore_data(self):
        shutil.copy(self.backup_file, self.excel_file)
        self.wb = load_workbook(self.excel_file)
        self.load_students_from_excel()
        self.load_books_from_excel()
        self.load_borrowing_from_excel()
        self.load_returned_from_excel()
        QMessageBox.information(self, "استعادة البيانات", "تم استعادة البيانات بنجاح.")

    def generate_report(self):
        ws_borrowing = self.wb['Borrowing']
        ws_returned = self.wb['Returned']

        report_data = []
        for row in ws_borrowing.iter_rows(min_row=2, values_only=True):
            report_data.append(row)
        for row in ws_returned.iter_rows(min_row=2, values_only=True):
            report_data.append(row)

        self.reports_table.setRowCount(len(report_data))
        for i, row in enumerate(report_data):
            for j, value in enumerate(row):
                self.reports_table.setItem(i, j, QTableWidgetItem(str(value)))

if __name__ == "__main__":
    app = QApplication(sys.argv)
    system = LibraryManagementSystem()
    sys.exit(app.exec_())
